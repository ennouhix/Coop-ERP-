"""
Tests du module reporting.

Vérifie non seulement le code HTTP, mais que les fichiers générés sont
RÉELLEMENT des PDF/XLSX valides et exploitables (pas juste une réponse
200 avec un contenu vide ou corrompu).
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase
from rest_framework_simplejwt.tokens import RefreshToken

from apps.billing import services as billing_services
from apps.catalog.models import Product, Unit
from apps.cooperatives.models import Cooperative
from apps.inventory import services as inventory_services
from apps.members.services import create_member
from apps.partners.models import Partner
from apps.purchases import services as purchase_services
from apps.sales import services as sales_services
from apps.warehouses.models import Warehouse

User = get_user_model()


class ReportingTestCase(APITestCase):
    def setUp(self) -> None:
        cache.clear()
        self.cooperative = Cooperative.objects.create(
            name="Coopérative Argane",
            slug="argane",
            ice="001234567000012",
            legal_name="Argane SARL",
        )
        self.admin = User.objects.create_user(
            username="admin",
            email="admin@argane.ma",
            password="MotDePasseSolide123",
            cooperative=self.cooperative,
            role="admin",
        )
        self.staff = User.objects.create_user(
            username="staff",
            email="staff@argane.ma",
            password="MotDePasseSolide123",
            cooperative=self.cooperative,
            role="staff",
        )

        self.unit = Unit.objects.create(
            cooperative=self.cooperative, name="Kilogramme", symbol="kg", unit_type="weight"
        )
        self.product = Product.objects.create(
            cooperative=self.cooperative,
            sku="PRD-00001",
            name={"fr": "Huile d'argane"},
            unit=self.unit,
        )
        self.warehouse = Warehouse.objects.create(
            cooperative=self.cooperative, code="WH-0001", name="Entrepôt", is_default=True
        )
        self.customer = Partner.objects.create(
            cooperative=self.cooperative,
            code="PART-0001",
            name="Épicerie Al Baraka",
            is_customer=True,
            is_supplier=False,
            phone_number="0612345678",
        )

    def _auth(self, user) -> None:  # noqa: ANN001
        token = RefreshToken.for_user(user).access_token
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")

    def _create_issued_invoice(self):
        invoice = billing_services.create_manual_invoice(
            cooperative=self.cooperative,
            customer=self.customer,
            lines=[
                {
                    "product": self.product,
                    "description": "",
                    "quantity": Decimal("10"),
                    "unit_price": Decimal("50.00"),
                }
            ],
            actor=self.admin,
            issue_date=date(2026, 7, 1),
        )
        billing_services.issue_invoice(invoice=invoice, actor=self.admin)
        return invoice

    # --- Facture PDF ---

    def test_invoice_pdf_is_valid_pdf_document(self) -> None:
        invoice = self._create_issued_invoice()
        self._auth(self.admin)
        url = reverse("reporting:invoice-pdf", args=[invoice.id])
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        content = b"".join(response.streaming_content) if response.streaming else response.content
        self.assertTrue(content.startswith(b"%PDF"))
        self.assertGreater(len(content), 500)  # un PDF vide/corrompu ferait quelques octets

    def test_staff_cannot_download_invoice_pdf(self) -> None:
        invoice = self._create_issued_invoice()
        self._auth(self.staff)
        url = reverse("reporting:invoice-pdf", args=[invoice.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_cannot_download_invoice_pdf_from_other_cooperative(self) -> None:
        other_coop = Cooperative.objects.create(name="Autre Coop", slug="autre")
        other_admin = User.objects.create_user(
            username="other",
            email="other@autre.ma",
            password="MotDePasseSolide123",
            cooperative=other_coop,
            role="admin",
        )
        invoice = self._create_issued_invoice()
        self._auth(other_admin)
        url = reverse("reporting:invoice-pdf", args=[invoice.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    # --- Export membres ---

    def test_members_export_contains_correct_rows(self) -> None:
        create_member(
            cooperative=self.cooperative,
            first_name="Ahmed",
            last_name="Ouazzani",
            phone_number="0699999999",
        )
        create_member(
            cooperative=self.cooperative,
            first_name="Fatima",
            last_name="Bennani",
            phone_number="0688888888",
        )

        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-members"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "N° Adhérent")
        self.assertEqual(ws.max_row, 3)  # 1 en-tête + 2 membres
        self.assertEqual(ws.cell(row=2, column=1).value, "ARG-0001")

    def test_staff_cannot_export_members(self) -> None:
        self._auth(self.staff)
        response = self.client.get(reverse("reporting:export-members"))
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    # --- Export mouvements de stock ---

    def test_stock_movements_export_contains_correct_data(self) -> None:
        inventory_services.record_stock_in(
            product=self.product, warehouse=self.warehouse, quantity=Decimal("50"), actor=self.admin
        )
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-stock-movements"))
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.max_row, 2)  # 1 en-tête + 1 mouvement
        self.assertEqual(ws.cell(row=2, column=4).value, "PRD-00001")
        self.assertEqual(ws.cell(row=2, column=7).value, 50.0)

    def test_stock_movements_export_filters_by_date(self) -> None:
        inventory_services.record_stock_in(
            product=self.product, warehouse=self.warehouse, quantity=Decimal("50"), actor=self.admin
        )
        self._auth(self.admin)
        far_future = "2099-01-01"
        response = self.client.get(
            reverse("reporting:export-stock-movements"), {"date_from": far_future}
        )
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.max_row, 1)  # seulement l'en-tête, rien dans la période

    # --- Export commandes de vente ---

    def test_sales_orders_export_contains_correct_data(self) -> None:
        sales_services.create_sales_order(
            cooperative=self.cooperative,
            customer=self.customer,
            warehouse=self.warehouse,
            lines=[
                {
                    "product": self.product,
                    "quantity_ordered": Decimal("5"),
                    "unit_price": Decimal("50.00"),
                }
            ],
            actor=self.admin,
            order_date=date(2026, 7, 1),
        )
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-sales-orders"))
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).value, "SO-00001")
        self.assertEqual(ws.cell(row=2, column=5).value, 250.0)

    # --- Nouveaux exports (achats, partenaires, factures, stock, journal) ---

    def _create_purchase_order(self):
        supplier = Partner.objects.create(
            cooperative=self.cooperative,
            code="SUP-0001",
            name="Fournisseur Atlas",
            is_supplier=True,
            phone_number="0611111111",
        )
        return purchase_services.create_purchase_order(
            cooperative=self.cooperative,
            supplier=supplier,
            warehouse=self.warehouse,
            lines=[
                {
                    "product": self.product,
                    "quantity_ordered": Decimal("10"),
                    "unit_price": Decimal("30.00"),
                }
            ],
            actor=self.admin,
            order_date=date(2026, 7, 10),
        )

    def test_purchase_orders_export_contains_correct_data(self) -> None:
        self._create_purchase_order()
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-purchase-orders"))
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=2).value, "Fournisseur")
        self.assertEqual(ws.cell(row=2, column=1).value, "PO-00001")
        self.assertEqual(ws.cell(row=2, column=2).value, "Fournisseur Atlas")
        self.assertEqual(ws.cell(row=2, column=5).value, 300.0)

    def test_partners_export_filters_by_kind(self) -> None:
        Partner.objects.create(
            cooperative=self.cooperative,
            code="SUP-0001",
            name="Fournisseur Atlas",
            is_supplier=True,
            phone_number="0622222222",
        )
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-partners"), {"kind": "supplier"})
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        # Le client de la fixture est exclu, seul le fournisseur apparaît.
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=2).value, "Fournisseur Atlas")

    def test_invoices_export_contains_correct_data(self) -> None:
        self._create_issued_invoice()
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-invoices"))
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "N° Facture")
        self.assertEqual(ws.cell(row=2, column=5).value, 500.0)
        self.assertEqual(ws.cell(row=2, column=7).value, 500.0)

    def test_stock_levels_export_contains_correct_data(self) -> None:
        inventory_services.record_stock_in(
            product=self.product, warehouse=self.warehouse, quantity=Decimal("50"), actor=self.admin
        )
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-stock-levels"))
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).value, "PRD-00001")
        self.assertEqual(ws.cell(row=2, column=4).value, 50.0)

    def _create_accounting_journal(self):
        from apps.accounting.models import Account, Journal

        journal = Journal.objects.create(
            cooperative=self.cooperative,
            code="JV",
            name={"fr": "Journal de vente"},
            journal_type="sales",
        )
        account = Account.objects.create(
            cooperative=self.cooperative,
            code="5141",
            name={"fr": "Banque"},
            account_type="treasury",
        )
        return journal, account

    def test_accounting_journal_export_contains_correct_data(self) -> None:
        journal, account = self._create_accounting_journal()
        from apps.accounting.services import create_accounting_entry

        create_accounting_entry(
            cooperative=self.cooperative,
            journal=journal,
            entry_date=date(2026, 7, 15),
            description="Vente n°1",
            lines_data=[
                {
                    "account": account,
                    "label": "Banque",
                    "debit": Decimal("100.00"),
                    "credit": Decimal("0"),
                },
                {
                    "account": account,
                    "label": "Ventes",
                    "debit": Decimal("0"),
                    "credit": Decimal("100.00"),
                },
            ],
            actor=self.admin,
        )
        self._auth(self.admin)
        response = self.client.get(
            reverse("reporting:export-accounting-journal"), {"period": "2026-07"}
        )
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.max_row, 3)  # 1 en-tête + 2 lignes d'écriture
        self.assertEqual(ws.cell(row=2, column=3).value, "JV")
        self.assertEqual(ws.cell(row=2, column=6).value, 100.0)

    def test_accounting_journal_export_filters_by_period(self) -> None:
        journal, account = self._create_accounting_journal()
        from apps.accounting.services import create_accounting_entry

        create_accounting_entry(
            cooperative=self.cooperative,
            journal=journal,
            entry_date=date(2026, 7, 15),
            lines_data=[
                {
                    "account": account,
                    "label": "Banque",
                    "debit": Decimal("100.00"),
                    "credit": Decimal("0"),
                },
                {
                    "account": account,
                    "label": "Ventes",
                    "debit": Decimal("0"),
                    "credit": Decimal("100.00"),
                },
            ],
            actor=self.admin,
        )
        self._auth(self.admin)
        response = self.client.get(
            reverse("reporting:export-accounting-journal"), {"period": "2099-01"}
        )
        content = b"".join(response.streaming_content) if response.streaming else response.content
        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb.active.max_row, 1)

    # --- Export PDF ---

    def test_members_export_pdf_is_valid_document(self) -> None:
        create_member(cooperative=self.cooperative, first_name="Ahmed", last_name="Ouazzani")
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:export-members"), {"output": "pdf"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        content = b"".join(response.streaming_content) if response.streaming else response.content
        self.assertTrue(content.startswith(b"%PDF"))
        self.assertGreater(len(content), 500)

    # --- Aperçu à l'écran (preview JSON) ---

    def test_preview_members_returns_columns_rows_and_total(self) -> None:
        create_member(cooperative=self.cooperative, first_name="Ahmed", last_name="Ouazzani")
        create_member(cooperative=self.cooperative, first_name="Fatima", last_name="Bennani")
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:preview-report", args=["members"]))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["report"], "members")
        self.assertEqual(payload["columns"][0], "N° Adhérent")
        self.assertEqual(payload["total"], 2)
        self.assertEqual(len(payload["rows"]), 2)
        self.assertFalse(payload["truncated"])

    def test_staff_cannot_access_preview(self) -> None:
        self._auth(self.staff)
        response = self.client.get(reverse("reporting:preview-report", args=["members"]))
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_preview_unknown_report_returns_404(self) -> None:
        self._auth(self.admin)
        response = self.client.get(reverse("reporting:preview-report", args=["inconnu"]))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
