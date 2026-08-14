"""
Exports Excel — un patron simple réutilisable : en-têtes stylés, largeurs
de colonnes ajustées, une feuille par export. Les lignes proviennent de
`data.py` (source unique) ; les `Decimal` sont écrits en numériques pour
rester sommables dans Excel.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.cooperatives.models import Cooperative
from apps.reporting import data

HEADER_FILL = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(ws, headers: list) -> None:  # noqa: ANN001
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)


def _write_rows(ws, rows: list) -> None:  # noqa: ANN001
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)


def from_rows(sheet_title: str, headers: list, rows: list) -> BytesIO:  # noqa: ANN001
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # limite openpyxl : 31 caractères max
    _write_header(ws, headers)
    _write_rows(ws, rows)
    return _to_buffer(wb)


def _to_buffer(workbook: Workbook) -> BytesIO:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# --- Adhérents ---

def export_members_excel(cooperative: Cooperative) -> BytesIO:
    headers, rows = data.members_rows(cooperative)
    return from_rows("Membres", headers, rows)


# --- Mouvements de stock ---

def export_stock_movements_excel(
    cooperative: Cooperative,
    date_from: date | None = None,
    date_to: date | None = None,
    movement_type: str | None = None,
    warehouse_id: str | None = None,
) -> BytesIO:
    headers, rows = data.stock_movements_rows(
        cooperative, date_from=date_from, date_to=date_to,
        movement_type=movement_type, warehouse_id=warehouse_id,
    )
    return from_rows("Mouvements de stock", headers, rows)


# --- Commandes de vente ---

def export_sales_orders_excel(
    cooperative: Cooperative,
    date_from: date | None = None,
    date_to: date | None = None,
    status: str | None = None,
    customer_id: str | None = None,
) -> BytesIO:
    headers, rows = data.sales_orders_rows(
        cooperative, date_from=date_from, date_to=date_to,
        status=status, customer_id=customer_id,
    )
    return from_rows("Commandes de vente", headers, rows)


# --- Commandes d'achat ---

def export_purchase_orders_excel(
    cooperative: Cooperative,
    date_from: date | None = None,
    date_to: date | None = None,
    status: str | None = None,
    supplier_id: str | None = None,
) -> BytesIO:
    headers, rows = data.purchase_orders_rows(
        cooperative, date_from=date_from, date_to=date_to,
        status=status, supplier_id=supplier_id,
    )
    return from_rows("Commandes d'achat", headers, rows)


# --- Partenaires ---

def export_partners_excel(
    cooperative: Cooperative,
    kind: str | None = None,
    status: str | None = None,
) -> BytesIO:
    headers, rows = data.partners_rows(cooperative, kind=kind, status=status)
    return from_rows("Partenaires", headers, rows)


# --- Factures ---

def export_invoices_excel(
    cooperative: Cooperative,
    date_from: date | None = None,
    date_to: date | None = None,
    status: str | None = None,
    customer_id: str | None = None,
) -> BytesIO:
    headers, rows = data.invoices_rows(
        cooperative, date_from=date_from, date_to=date_to,
        status=status, customer_id=customer_id,
    )
    return from_rows("Factures", headers, rows)


# --- Niveaux de stock ---

def export_stock_levels_excel(
    cooperative: Cooperative,
    warehouse_id: str | None = None,
) -> BytesIO:
    headers, rows = data.stock_levels_rows(cooperative, warehouse_id=warehouse_id)
    return from_rows("Niveaux de stock", headers, rows)


# --- Journal comptable ---

def export_accounting_journal_excel(
    cooperative: Cooperative,
    period: str | None = None,
    journal_id: str | None = None,
) -> BytesIO:
    headers, rows = data.accounting_journal_rows(
        cooperative, period=period, journal_id=journal_id,
    )
    return from_rows("Journal comptable", headers, rows)
